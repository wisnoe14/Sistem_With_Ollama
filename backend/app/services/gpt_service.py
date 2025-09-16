import hashlib
import os
import json
import re
import requests
import subprocess
from typing import Union


def truncate_to_n_words(text, n=7):
    """Potong teks maksimal n kata."""
    return " ".join(text.split()[:n])


# Cache dictionary untuk response Ollama
_ollama_cache = {}


def cache_key(*args, **kwargs):
    key_str = str(args) + str(kwargs)
    return hashlib.md5(key_str.encode()).hexdigest()


OLLAMA_URL = os.getenv("OLLAMA_URL", "http://localhost:11434/api/chat")
OLLAMA_MODEL = os.getenv("OLLAMA_MODEL", "llama3")
1

# Cek dan pull model jika belum ada
import requests as reqs
def ensure_ollama_model(model_name: str):
    try:
        resp = reqs.get("http://localhost:11434/api/tags")
        if resp.status_code == 200:
            tags = resp.json().get("models", [])
            if not any(m.get("name", "").startswith(model_name) for m in tags):
                print(f"[INFO] Model {model_name} belum ada, melakukan pull...")
                subprocess.run(["ollama", "pull", model_name], check=True)
    except Exception as e:
        print(f"[WARNING] Tidak bisa cek/pull model Ollama: {e}")


ensure_ollama_model(OLLAMA_MODEL)


def predict_status_promo_ollama(answers: list) -> dict:
    """
    Prediksi status, promo, estimasi bayar, alasan, dan minat berlangganan menggunakan Ollama.
    """
    percakapan = " | ".join([str(a) for a in answers if a])
    prompt = (
        f"Percakapan: {percakapan}. "
        f"Prediksi status (Pelanggan tidak dapat dihubungi, Closing, Pelanggan dapat dihubungi, Bersedia Membayar), "
        f"promo (Tidak Ada Promo, Promo Diskon, Promo Cashback, Promo Gratis Bulan, Promo Lainnya), "
        f"minat berlangganan (Ya, Tidak, Ragu). "
        f"Format: Status: <status>, Promo: <jenis_promo>, Estimasi: <estimasi atau 'Belum tersedia'>, "
        f"Alasan: <alasan ringkas, max 7 kata>, Minat: <Ya/Tidak/Ragu>."
    )

    key = cache_key(prompt, "predict_status_promo_ollama_chat")
    if key in _ollama_cache:
        content = _ollama_cache[key]
    else:
        payload = {
            "model": OLLAMA_MODEL,
            "messages": [
                {
                    "role": "system",
                    "content": "Kamu adalah asisten customer service yang membantu memprediksi status, promo, dan minat pelanggan."
                },
                {"role": "user", "content": prompt}
            ],
            "stream": False
        }
        try:
            resp = requests.post(OLLAMA_URL, json=payload, timeout=120)
            resp.raise_for_status()
            data = resp.json()
            content = data.get("message", {}).get("content", "")
        except Exception as e:
            print(f"[OLLAMA ERROR][predict_status_promo_ollama] {e}")
            content = ""
        _ollama_cache[key] = content

    # Parsing output dengan regex
    result = {}
    patterns = {
        'status': r'Status\s*:\s*([^,\n]+)',
        'promo': r'Promo\s*:\s*([^,\n]+)',
        'estimasi': r'Estimasi\s*:\s*([^,\n]+)',
        'alasan': r'Alasan\s*:\s*([^,\n]+)',
        'minat': r'Minat\s*:\s*([^,\n]+)'
    }
    for key, pat in patterns.items():
        m = re.search(pat, content, re.IGNORECASE)
        if m:
            result[key] = m.group(1).strip()
        else:
            result[key] = ""

    return {
        "status": result.get("status", ""),
        "jenis_promo": result.get("promo", ""),
        "estimasi_pembayaran": result.get("estimasi", ""),
        "alasan": result.get("alasan", ""),
        "minat_berlangganan": result.get("minat", "")
    }


def generate_question(topic: str, context: Union[str, list] = "") -> dict:
    """
    Generate pertanyaan + opsi jawaban dari context percakapan (via /api/chat).
    Hanya relevan dengan domain ICONNET (internet, tagihan, promo, retention, winback).
    """
    percakapan = ""
    if isinstance(context, list):
        percakapan = " | ".join([
            f"Q:{c.get('q','')} A:{c.get('a','')}"
            for c in context if isinstance(c, dict) and c.get('a', '').strip()
        ])
    elif isinstance(context, str):
        percakapan = context

    last_answer = ""
    if isinstance(context, list) and len(context) > 0:
        last_answer = context[-1].get('a', '')

    prompt = f"""
"Sebagai Customer Service ICONNET, buatlah pertanyaan yang sesuai dengan jawaban pelanggan sebelumnya berkaitan dengan topik {topic}. Pertanyaan harus singkat, formal, dan sopan. Sertakan 4 opsi jawaban singkat yang relevan dengan layanan ICONNET yang mencakup internet, promo, pembayaran, retention, atau winback. Pastikan pertanyaan tidak bersifat umum dan fokus pada layanan internet."

Contoh Format:
{{
  "question": "Apa alasan utama Anda menunda pembayaran?",
  "options": ["Belum gajian", "Lupa bayar", "Tagihan terlalu tinggi", "Lainnya"]
}}

Percakapan sejauh ini: {percakapan}
Jawaban terakhir pelanggan: {last_answer}

Buat pertanyaan baru dengan format JSON seperti contoh.
""".strip()

    key = cache_key(prompt, "generate_question_ollama_chat")
    if key in _ollama_cache:
        result_json = _ollama_cache[key]
    else:
        payload = {
            "model": OLLAMA_MODEL,
            "messages": [
                {"role": "system", "content": "Kamu adalah CS ICONNET yang hanya menanyakan hal terkait layanan internet ICONNET."},
                {"role": "user", "content": prompt}
            ],
            "stream": False
        }
        try:
            resp = requests.post(OLLAMA_URL, json=payload, timeout=120)
            resp.raise_for_status()
            data = resp.json()
            result_json = data.get("message", {}).get("content", "")
        except Exception as e:
            print(f"[OLLAMA ERROR][generate_question] {e}")
            result_json = ""
        _ollama_cache[key] = result_json

    # Parsing JSON hasil model
    try:
        match = re.search(r"\{.*\}", result_json, re.DOTALL)
        if match:
            data = json.loads(match.group(0))
        else:
            data = {}

        q = truncate_to_n_words(data.get("question", ""), 12)
        opts = [truncate_to_n_words(o, 7) for o in data.get("options", []) if isinstance(o, str) and o.strip()]

        # Fallback jika hasil kosong / tidak relevan
        if not q or not opts:
            if topic == "retention":
                q = "Apa kendala utama Anda dalam menggunakan ICONNET?"
                opts = ["Tagihan mahal", "Gangguan teknis", "Layanan kurang", "Lainnya"]
            elif topic == "winback":
                q = "Apa alasan Anda ingin berhenti berlangganan?"
                opts = ["Harga mahal", "Kualitas layanan", "Butuh promo", "Lainnya"]
            else:
                q = "Apa alasan utama keterlambatan pembayaran ICONNET?"
                opts = ["Belum gajian", "Lupa bayar", "Tagihan tinggi", "Lainnya"]

        return {"question": q, "options": opts}

    except Exception as e:
        print("[DEBUG PARSE ERROR]", e)
        return {"question": "Pertanyaan tidak tersedia", "options": ["Jawaban 1", "Jawaban 2", "Jawaban 3", "Jawaban 4"]}


def save_conversation_to_excel(
    customer_id: str,
    mode: str,
    status_dihubungi: str,
    percakapan: list,
    prediction: dict = None,
    filename: str = "riwayat_simulasi_cs.xlsx"
) -> str:
    """
    Simpan riwayat percakapan ke file Excel (openpyxl).
    Kolom: customer_id, mode, status_dihubungi, pertanyaan, jawaban,
           prediksi_status, promo, estimasi_bayar, alasan, minat_berlangganan, timestamp
    """
    from openpyxl import Workbook, load_workbook
    import datetime

    if prediction is None:
        answers = [item.get("a", "") for item in percakapan if item.get("a", "")]
        prediction = predict_status_promo_ollama(answers)

    pred_status = prediction.get("status", "")
    promo = prediction.get("jenis_promo", "")
    estimasi_bayar = prediction.get("estimasi_pembayaran", "")
    alasan = prediction.get("alasan", "")
    minat = prediction.get("minat_berlangganan", "")
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append([
            "customer_id", "mode", "status_dihubungi", "pertanyaan", "jawaban",
            "prediksi_status", "promo", "estimasi_bayar", "alasan", "minat_berlangganan", "timestamp"
        ])

    for step in percakapan:
        pertanyaan = step.get("q", "")
        jawaban = step.get("a", "")
        ws.append([
            customer_id,
            mode,
            status_dihubungi,
            pertanyaan,
            jawaban,
            pred_status,
            promo,
            estimasi_bayar,
            alasan,
            minat,
            timestamp
        ])
    wb.save(filename)
    return filename
